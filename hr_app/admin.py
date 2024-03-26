from django.contrib import admin
from django.http import HttpResponse
from .models import registration,Review
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime


class registrationAdmin(admin.ModelAdmin):
    list_display = ('created_at',"Name","Primary_contact",'Secondary_contact',"Location",'Email_Id','Soft_Skills','Technical_Skills','General_Skills','Current_CTC','Expected_CTC','Notice_Period','Current_designation','Applied_designation','experience','Job_portal_source','Contacted_by')  
    
    search_fields = ("Soft_Skills",'Technical_Skills','General_Skills','Name','Email_Id')  
    list_filter = ('Email_Id',) 
    
    
    actions = ['download_selected']


    def download_selected(self, request, queryset):
        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Define the header
        header = ["Name", "Primary_contact", "Secondary_contact", "Location", "Email_Id", "Soft_Skills", "Technical_Skills", 
                "General_Skills", "Current_CTC", "Expected_CTC", "Notice_period", "Current_designation", 
                "Applied_designation", "experience", "Job_portal_source", "Contacted_by"]

        # Add header to the worksheet
        for col_num, header_text in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}1"]
            cell.value = header_text
            cell.alignment = Alignment(horizontal='center')

        # Add data to the worksheet
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(header, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}{row_num}"]
                cell.value = getattr(obj, field_name, None)
                cell.alignment = Alignment(horizontal='center')

        # Create the HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        response['Content-Disposition'] = f'attachment; filename= "{current_date} .xlsx"'
        workbook.save(response)

        return response


    download_selected.short_description = "Download selected items"

# Register the model with the custom admin class
admin.site.register(registration, registrationAdmin)
   

class ReviewAdmin(admin.ModelAdmin):
    list_display=(
    'user_name',
    'Moral_character',
    'punctuality',
    'health',
    'behaviour',
    'attitude',
    'Career_goals',
    'understanding_level',
    'possitive_attitude_and_mind',
    'executive',
    'responsibility',
    'response_ability',
    'team_handling',
    'planing',
    'communicate_ability',
    'passion',
    'confidence',
    'profissional_background',
    'work_experience',
    'knowledge_level',
    'english_skils',
    'other_languages',
    'consider_to_client',
    'Internal_hiring',
    'reject',
    'created_at',
    'remarks',
)
    list_filter = ('Internal_hiring','reject') 
    #autocomplete_fields = ['user_name'] 
    autocomplete_fields = ['user_name']
    
    

admin.site.register(Review, ReviewAdmin)
admin.site.site_header=" HR Adminstration"

